import os
import sys
import datetime

sys.stdout.reconfigure(encoding='utf-8')

from fastapi import FastAPI, Depends, HTTPException, WebSocket, WebSocketDisconnect, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import pandas as pd
import openpyxl
import io

import models
import database
import hooks
import ai_agent

# Initialize FastAPI App
app = FastAPI(title="Real-time Task & Budget Management System")

# CORS Middleware (allows Streamlit frontend to interact with the API)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Add No-Cache Middleware for frontend updates
@app.middleware("http")
async def add_no_cache_headers(request, call_next):
    response = await call_next(request)
    path = request.url.path.lower()
    if path == "/" or path.endswith(".html") or path.endswith(".js") or path.endswith(".css") or "/static/" in path:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response

# Pydantic schemas for request validation
class TaskProgressUpdate(BaseModel):
    tien_do: float
    trang_thai: str
    dieu_kien_ghi_nhan: Optional[str] = ""

class SpendingCreate(BaseModel):
    task_id: int
    so_tien_chi: float
    nguoi_cap_nhat: str
    chung_tu_kem_theo: Optional[str] = ""
    trang_thai_duyet: Optional[str] = "Approved"

class AIUpdateInput(BaseModel):
    text: str
    api_key: Optional[str] = None

# WebSocket Hub
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        print(f"WebSocket client connected. Total: {len(self.active_connections)}")

    def disconnect(self, websocket: WebSocket):
        if websocket in self.active_connections:
            self.active_connections.remove(websocket)
            print(f"WebSocket client disconnected. Total: {len(self.active_connections)}")

    async def broadcast(self, message: dict):
        for connection in self.active_connections:
            try:
                await connection.send_json(message)
            except Exception as e:
                # Connection might be closed, we will clean it up later or handle gracefully
                print(f"Error sending message to WebSocket client: {e}")

manager = ConnectionManager()

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            # Keep connection alive, listen for messages if needed
            data = await websocket.receive_text()
            # Echo back or ignore client messages
            await websocket.send_json({"type": "ping", "message": "alive"})
    except WebSocketDisconnect:
        manager.disconnect(websocket)
    except Exception as e:
        print(f"WebSocket error: {e}")
        manager.disconnect(websocket)

# ----------------- API Endpoints -----------------

import re
import calendar

def parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    
    # 2026-06-30
    match1 = re.match(r'^(\d{4})-(\d{2})-(\d{2})$', date_str)
    if match1:
        try:
            return datetime.date(int(match1.group(1)), int(match1.group(2)), int(match1.group(3)))
        except ValueError:
            pass
            
    # Tháng 06/2026
    match2 = re.match(r'^(?:Tháng|tháng|Th\s*)\s*(\d{2})/(\d{4})$', date_str)
    if match2:
        try:
            month = int(match2.group(1))
            year = int(match2.group(2))
            last_day = calendar.monthrange(year, month)[1]
            return datetime.date(year, month, last_day)
        except Exception:
            pass
            
    # 06/2026
    match3 = re.match(r'^(\d{2})/(\d{4})$', date_str)
    if match3:
        try:
            month = int(match3.group(1))
            year = int(match3.group(2))
            last_day = calendar.monthrange(year, month)[1]
            return datetime.date(year, month, last_day)
        except Exception:
            pass
            
    return None

def format_date(d, original_str=None):
    if d:
        return d.strftime("%Y-%m-%d")
    if original_str:
        parsed = parse_date(original_str)
        if parsed:
            return parsed.strftime("%Y-%m-%d")
    return ""

def compute_rolled_up_deadlines(all_tasks):
    computed = {}
    
    # 1. Parse all tasks' deadlines first
    for t in all_tasks:
        computed[t.stt] = parse_date(t.thoi_han_hoan_thanh)
        
    # 2. Roll up hierarchy (post-order-like traversal via prefix check)
    for t in all_tasks:
        descendant_dates = []
        for other in all_tasks:
            # Check if other is a descendant of t (starts with t.stt + ".")
            if other.stt.startswith(t.stt + "."):
                d_date = parse_date(other.thoi_han_hoan_thanh)
                if d_date:
                    descendant_dates.append(d_date)
                    
        if descendant_dates:
            # Add own date if it exists
            own_date = parse_date(t.thoi_han_hoan_thanh)
            if own_date:
                descendant_dates.append(own_date)
            max_date = max(descendant_dates)
            computed[t.stt] = max_date
            
    return computed

def get_corresponding_tkbvtc_deadline_stt(task_name):
    name_upper = task_name.upper()
    if "KV 2" in name_upper or "KV2" in name_upper:
        return "14.3.1"
    elif "3.1" in name_upper:
        return "14.3.2"
    elif "3.2" in name_upper:
        return "14.3.3"
    elif "3.3" in name_upper:
        return "14.3.4"
    elif "3.4" in name_upper:
        return "14.3.5"
    elif "KV 4" in name_upper or "KV4" in name_upper:
        return "14.3.6"
    elif "CAO TẦNG" in name_upper or "NCT" in name_upper:
        return "14.3.7"
    elif "TDTT" in name_upper or "THỂ DỤC THỂ THAO" in name_upper:
        return "14.4.1"
    elif "VUI CHƠI" in name_upper or "TRẺ EM" in name_upper:
        return "14.4.2"
    return None

def calculate_phase2_status(task, rolled_up_deadlines):
    # 4. Hoàn thành: Khi đã phê duyệt hoàn thành hoặc tiến độ >= 100%
    if task.tien_do >= 100 or task.trang_thai == "Done":
        return "Done"

    current_date = datetime.date(2026, 7, 7) # Simulation today date
    comp_date = parse_date(task.thoi_han_hoan_thanh)
    
    # Tìm thời hạn phê duyệt HS TKBVTC tương ứng
    tkbvtc_stt = get_corresponding_tkbvtc_deadline_stt(task.ten_cong_viec)
    tkbvtc_date = rolled_up_deadlines.get(tkbvtc_stt) if tkbvtc_stt else None
    if not tkbvtc_date:
        tkbvtc_date = datetime.date(2026, 6, 30) # Mặc định làm mốc dự phòng
        
    if not comp_date:
        return task.trang_thai

    # 1. Chưa thực hiện: Thời gian thực < thời hạn phê duyệt HS TKBVTC
    if current_date < tkbvtc_date:
        return "Todo"

    # 3. Chậm: Thời gian thực > Thời hạn hoàn thành - 20 ngày (bao gồm cả khi vượt quá thời hạn hoàn thành)
    if current_date > (comp_date - datetime.timedelta(days=20)):
        return "Delayed"

    # 2. Đang thực hiện: thời hạn phê duyệt HS TKBVTC <= Thời gian thực <= thời hạn hoàn thành
    if tkbvtc_date <= current_date <= comp_date:
        return "In-Progress"

    return task.trang_thai

def calculate_phase134_status(task, rolled_up_deadlines):
    # Hoàn thành: tiến độ >= 100% hoặc Done
    if task.tien_do >= 100 or task.trang_thai == "Done":
        return "Done"
        
    # Lấy thời hạn (đã gộp từ con hoặc gốc)
    rolled_date = rolled_up_deadlines.get(task.stt)
    if not rolled_date:
        rolled_date = parse_date(task.thoi_han_hoan_thanh)
        
    if not rolled_date:
        # Nếu không có thời hạn hoàn thành, giữ nguyên trạng thái
        return task.trang_thai
        
    current_date = datetime.date(2026, 7, 7) # Simulation today date
    
    # Nếu thời gian thực < thời hạn thì Đang thực hiện
    if current_date < rolled_date:
        return "In-Progress"
    else:
        # Nếu thời gian thực >= thời hạn thì Quá hạn
        return "Delayed"

def get_vietnamese_status(status_str):
    if not status_str:
        return "Chưa thực hiện"
    status_lower = str(status_str).lower()
    if "todo" in status_lower:
        return "Chưa thực hiện"
    elif "in-progress" in status_lower or "in_progress" in status_lower:
        return "Đang thực hiện"
    elif "delayed" in status_lower or "chậm" in status_lower or "trễ" in status_lower:
        return "Quá hạn"
    elif "done" in status_lower or "hoàn thành" in status_lower:
        return "Hoàn thành"
    return status_str

def log_action(db: Session, username: Optional[str], action: str, details: str):
    ho_ten = "Khách vãng lai"
    user_name = "guest"
    if username:
        user = db.query(models.User).filter(models.User.username == username).first()
        if user:
            ho_ten = user.ho_ten
            user_name = user.username
        else:
            user_name = username
            ho_ten = username
            
    log_entry = models.ActionLog(
        username=user_name,
        ho_ten=ho_ten,
        hanh_dong=action,
        chi_tiet=details,
        thoi_gian=datetime.datetime.now()
    )
    db.add(log_entry)
    db.flush()

@app.get("/api/logs")
def get_action_logs(db: Session = Depends(database.get_db)):
    logs = db.query(models.ActionLog).order_by(models.ActionLog.id.desc()).limit(200).all()
    return [{
        "id": l.id,
        "username": l.username,
        "ho_ten": l.ho_ten,
        "hanh_dong": l.hanh_dong,
        "chi_tiet": l.chi_tiet,
        "thoi_gian": l.thoi_gian.strftime("%Y-%m-%d %H:%M:%S")
    } for l in logs]

def check_task_update_permissions(user, task, request_data, is_partial=False):
    # Rule 0: Lock after weekly approval. Non-admins cannot edit if duyet_tuan is "Đã duyệt"
    if task.duyet_tuan == "Đã duyệt":
        if not user or user.role != "Admin":
            raise HTTPException(
                status_code=403,
                detail="Công việc đã được phê duyệt tuần. Không thể chỉnh sửa thêm."
            )

    # Check task level dynamically based on STT formatting
    task_level = task.stt.count('.') + 1
    
    # Check if there are structural changes in the request
    has_structural_changes = False
    if request_data.ma_ngan_sach != task.ma_ngan_sach: has_structural_changes = True
    if request_data.ten_cong_viec != task.ten_cong_viec: has_structural_changes = True
    if request_data.phong_ban_thuc_hien != task.phong_ban_thuc_hien: has_structural_changes = True
    if request_data.co_quan_giai_quyet != task.co_quan_giai_quyet: has_structural_changes = True
    if request_data.ho_so_dau_ra != task.ho_so_dau_ra: has_structural_changes = True
    if request_data.dieu_kien_ghi_nhan != task.dieu_kien_ghi_nhan: has_structural_changes = True
    if request_data.thoi_han_hoan_thanh != task.thoi_han_hoan_thanh: has_structural_changes = True
    
    # Check if there are weekly CBQL fields changed (cach_giai_quyet or duyet_tuan)
    has_manager_changes = False
    if request_data.cach_giai_quyet != task.cach_giai_quyet: has_manager_changes = True
    if request_data.duyet_tuan != task.duyet_tuan: has_manager_changes = True

    # If no user is logged in, restrict
    if not user:
        raise HTTPException(status_code=403, detail="Yêu cầu đăng nhập để chỉnh sửa.")

    user_role = user.role
    user_dept = str(user.phong_ban).upper().strip()
    task_dept = str(task.phong_ban_thuc_hien).upper().strip()
    is_same_dept = user_dept == task_dept or user_dept == "ALL"

    # Admin has all permissions on all levels
    if user_role == "Admin":
        return

    # Check department membership first (all non-admin roles must belong to the task's department)
    if not is_same_dept:
        raise HTTPException(
            status_code=403,
            detail=f"Quyền hạn bị từ chối: Phòng ban của bạn ({user.phong_ban}) không trùng khớp với phòng thực hiện công việc ({task.phong_ban_thuc_hien})."
        )

    # Check if the user is the original creator/assignee of the task
    user_fullname = str(user.ho_ten).lower().strip()
    user_username = str(user.username).lower().strip()
    task_pt = str(task.nguoi_phu_trach).lower().strip()
    task_bc = str(task.nguoi_bao_cao).lower().strip()
    is_creator = user_fullname in (task_pt, task_bc) or user_username in (task_pt, task_bc)

    # 1. Structural fields changes restriction: Only Admin (already checked) or Task Creator can modify WBS structure, roles, timeline
    if has_structural_changes:
        if not is_creator:
            raise HTTPException(
                status_code=403,
                detail="Quyền hạn bị từ chối: Chỉ có Admin hoặc người lập ra công việc này mới được chỉnh sửa cấu trúc (Thông tin WBS, Nhân sự, Mốc thời gian)."
            )

    # 2. CBQL weekly approval changes restriction: Only Admin (already checked) or Trưởng phòng/CBQL can approve
    if has_manager_changes:
        if user_role not in ("TruongPhong", "CBQL", "PM"):
            raise HTTPException(
                status_code=403,
                detail="Quyền hạn bị từ chối: Nhân viên không có quyền phê duyệt tuần hoặc cập nhật ý kiến của quản lý."
            )
        return

class UserCreate(BaseModel):
    username: str
    ho_ten: str
    phong_ban: str
    role: str
    password: Optional[str] = "123456"

class LoginRequest(BaseModel):
    username: str
    password: str

@app.post("/api/auth/login")
def login_auth(request: LoginRequest, db: Session = Depends(database.get_db)):
    user = db.query(models.User).filter(models.User.username == request.username).first()
    if not user:
        raise HTTPException(status_code=400, detail="Tài khoản không tồn tại!")
    if user.password != request.password:
        raise HTTPException(status_code=400, detail="Mật khẩu không chính xác!")
    
    log_action(db, user.username, "Đăng nhập", f"Nhân sự {user.ho_ten} đăng nhập thành công.")
    
    return {
        "status": "success",
        "user": {
            "id": user.id,
            "username": user.username,
            "ho_ten": user.ho_ten,
            "phong_ban": user.phong_ban,
            "role": user.role
        }
    }

class ChangePasswordRequest(BaseModel):
    username: str
    old_password: str
    new_password: str

@app.put("/api/users/change-password")
def change_password(request: ChangePasswordRequest, db: Session = Depends(database.get_db)):
    user = db.query(models.User).filter(models.User.username == request.username).first()
    if not user:
        raise HTTPException(status_code=404, detail="Tài khoản không tồn tại!")
    
    if user.password != request.old_password:
        raise HTTPException(status_code=400, detail="Mật khẩu cũ không chính xác!")
    
    user.password = request.new_password
    db.add(user)
    log_action(db, user.username, "Đổi mật khẩu", f"Nhân sự {user.ho_ten} đã thay đổi mật khẩu thành công.")
    db.commit()
    return {"status": "success", "message": "Thay đổi mật khẩu thành công!"}

@app.get("/api/users")
def get_users(db: Session = Depends(database.get_db)):
    users = db.query(models.User).all()
    return [{"id": u.id, "username": u.username, "ho_ten": u.ho_ten, "phong_ban": u.phong_ban, "role": u.role, "password": u.password} for u in users]

@app.post("/api/users")
def create_user(
    user_data: UserCreate, 
    admin_username: Optional[str] = None, 
    db: Session = Depends(database.get_db)
):
    if not admin_username:
        raise HTTPException(status_code=403, detail="Yêu cầu đăng nhập admin để thực hiện.")
    admin = db.query(models.User).filter(models.User.username == admin_username).first()
    if not admin or admin.role != "Admin":
        raise HTTPException(status_code=403, detail="Chỉ có Admin hệ thống mới được phép thêm nhân sự.")
        
    existing = db.query(models.User).filter(models.User.username == user_data.username).first()
    if existing:
        raise HTTPException(status_code=400, detail="Tên đăng nhập đã tồn tại.")
        
    new_user = models.User(
        username=user_data.username.strip(),
        ho_ten=user_data.ho_ten.strip(),
        phong_ban=user_data.phong_ban.strip(),
        role=user_data.role.strip(),
        password=user_data.password.strip() if user_data.password else "123456"
    )
    db.add(new_user)
    db.commit()
    return {"status": "success", "user": {"id": new_user.id, "username": new_user.username}}

@app.put("/api/users/{user_id}")
def update_user(
    user_id: int, 
    user_data: UserCreate, 
    admin_username: Optional[str] = None, 
    db: Session = Depends(database.get_db)
):
    if not admin_username:
        raise HTTPException(status_code=403, detail="Yeu cau dang nhap admin de thuc hien.")
    admin = db.query(models.User).filter(models.User.username == admin_username).first()
    if not admin or admin.role != "Admin":
        raise HTTPException(status_code=403, detail="Chi co Admin he thong moi duoc phep sua nhan su.")
        
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Khong tim thay nhan su.")
        
    if user.username != user_data.username:
        existing = db.query(models.User).filter(models.User.username == user_data.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="Ten dang nhap da ton tai.")
            
    user.username = user_data.username.strip()
    user.ho_ten = user_data.ho_ten.strip()
    user.phong_ban = user_data.phong_ban.strip()
    user.role = user_data.role.strip()
    if user_data.password:
        user.password = user_data.password.strip()
    db.add(user)
    db.commit()
    return {"status": "success"}

@app.delete("/api/users/{user_id}")
def delete_user(
    user_id: int, 
    admin_username: Optional[str] = None, 
    db: Session = Depends(database.get_db)
):
    if not admin_username:
        raise HTTPException(status_code=403, detail="Yeu cau dang nhap admin de thuc hien.")
    admin = db.query(models.User).filter(models.User.username == admin_username).first()
    if not admin or admin.role != "Admin":
        raise HTTPException(status_code=403, detail="Chi co Admin he thong moi duoc phep xoa nhan su.")
        
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Khong tim thay nhan su.")
        
    if user.username == admin_username:
        raise HTTPException(status_code=400, detail="Khong duoc tu xoa tai khoan dang dang nhap.")
        
    db.delete(user)
    db.commit()
    return {"status": "success"}

@app.get("/api/project")
def get_project_info(db: Session = Depends(database.get_db)):
    project = db.query(models.Project).filter(models.Project.id == 1).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return {
        "ten_du_an": project.ten_du_an,
        "tong_ngan_sach": project.tong_ngan_sach
    }

@app.get("/api/phases")
def get_phases(db: Session = Depends(database.get_db)):
    phases = db.query(models.Phase).all()
    return phases

@app.get("/api/tasks")
def get_tasks(
    phase_id: Optional[int] = None,
    level: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    # Fetch all tasks to compute hierarchy rollups and cross-phase lookups properly
    all_tasks = db.query(models.Task).all()
    
    # Compute rolled up deadlines
    rolled_up_deadlines = compute_rolled_up_deadlines(all_tasks)
    
    # Filter by phase in memory
    filtered_tasks = all_tasks
    if phase_id:
        filtered_tasks = [t for t in all_tasks if t.phase_id == phase_id]
    
    # Now build the filtered list
    result = []
    for t in filtered_tasks:
        dot_count = t.stt.count('.')
        t_level = dot_count + 1
        
        # Level filter
        if level is not None and t_level != level:
            continue
            
        # Search filter
        if search:
            s_lower = search.lower()
            in_name = s_lower in (t.ten_cong_viec or "").lower()
            in_wbs = s_lower in (t.ma_ngan_sach or "").lower()
            in_stt = s_lower in (t.stt or "").lower()
            if not (in_name or in_wbs or in_stt):
                continue
                
        # Get rolled up deadline or keep original if none
        rolled_date = rolled_up_deadlines.get(t.stt)
        final_deadline = format_date(rolled_date, t.thoi_han_hoan_thanh)
        
        # Get computed status
        final_status = t.trang_thai
        if t.phase_id in (1, 3, 4):
            final_status = calculate_phase134_status(t, rolled_up_deadlines)
        elif t.phase_id == 2 and t_level == 3:
            final_status = calculate_phase2_status(t, rolled_up_deadlines)
        
        result.append({
            "id": t.id,
            "ma_ngan_sach": t.ma_ngan_sach,
            "stt": t.stt,
            "level": t_level,
            "phase_id": t.phase_id,
            "ten_cong_viec": t.ten_cong_viec,
            "phong_ban_thuc_hien": t.phong_ban_thuc_hien,
            "co_quan_giai_quyet": t.co_quan_giai_quyet,
            "ho_so_dau_ra": t.ho_so_dau_ra,
            "dieu_kien_ghi_nhan": t.dieu_kien_ghi_nhan,
            "thoi_han_hoan_thanh": final_deadline,
            "tien_do": t.tien_do,
            "trang_thai": final_status,
            "ke_hoach_tuan": t.ke_hoach_tuan,
            "ket_qua_tuan": t.ket_qua_tuan,
            "vuong_mac_tuan": t.vuong_mac_tuan,
            "cach_giai_quyet": t.cach_giai_quyet,
            "duyet_tuan": t.duyet_tuan,
            "ngay_khoi_tao": t.ngay_khoi_tao,
            "cong_trinh": t.cong_trinh,
            "doi_tac": t.doi_tac,
            "so_dien_thoai": t.so_dien_thoai,
            "ngay_bat_dau": t.ngay_bat_dau,
            "ngay_ket_thuc": t.ngay_ket_thuc,
            "gia_han_den_ngay": t.gia_han_den_ngay,
            "thoi_gian_bao_hanh": t.thoi_gian_bao_hanh,
            "mo_ta": t.mo_ta,
            "dieu_khoan": t.dieu_khoan,
            "nguoi_phu_trach": t.nguoi_phu_trach,
            "nguoi_bao_cao": t.nguoi_bao_cao,
            "nguoi_duyet": t.nguoi_duyet,
            "gia_tri_quyet_toan": t.gia_tri_quyet_toan,
            "da_nghiem_thu": t.da_nghiem_thu,
            "da_thanh_toan": t.da_thanh_toan,
            "tam_ung": t.tam_ung,
            "da_thu_hoi_tam_ung": t.da_thu_hoi_tam_ung,
            "weekly_reports_json": t.weekly_reports_json,
            "budget": {
                "ngan_sach_tong": t.budget.ngan_sach_tong if t.budget else 0.0,
                "is_locked": t.budget.is_locked if t.budget else False
            } if t.budget else None
        })
        
    return result

@app.get("/api/tasks/export")
def export_tasks_to_excel(db: Session = Depends(database.get_db)):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import DataBarRule
    
    tasks = db.query(models.Task).all()
    rolled_up_deadlines = compute_rolled_up_deadlines(tasks)
    
    # Sort them by STT hierarchically
    tasks = sorted(tasks, key=lambda x: [int(i) if i.isdigit() else 999 for i in x.stt.split('.')])
    
    rows = []
    for t in tasks:
        dot_count = t.stt.count('.')
        t_level = dot_count + 1
        budget_val = t.budget.ngan_sach_tong if t.budget else 0.0
        
        # Indent task name to represent level visually in Excel
        indented_name = ("    " * (t_level - 1)) + t.ten_cong_viec
        
        # Get rolled up deadline or keep original if none
        rolled_date = rolled_up_deadlines.get(t.stt)
        final_deadline = format_date(rolled_date, t.thoi_han_hoan_thanh)
        
        # Get computed status
        final_status = t.trang_thai
        if t.phase_id in (1, 3, 4):
            final_status = calculate_phase134_status(t, rolled_up_deadlines)
        elif t.phase_id == 2 and t_level == 3:
            final_status = calculate_phase2_status(t, rolled_up_deadlines)
            
        final_status_vn = get_vietnamese_status(final_status)
            
        rows.append({
            "STT": t.stt,
            "Cấp": t_level,
            "Mã Ngân Sách (WBS)": t.ma_ngan_sach,
            "Nội dung công việc": indented_name,
            "Phòng ban thực hiện": t.phong_ban_thuc_hien or "-",
            "Hồ sơ đầu ra": t.ho_so_dau_ra or "-",
            "Thời hạn hoàn thành": final_deadline or "-",
            "Điều kiện ghi nhận kết quả": t.dieu_kien_ghi_nhan or "-",
            "Ngân sách tổng (Trđ)": budget_val,
            "Tiến độ (%)": (t.tien_do / 100.0) if t.tien_do else 0.0,
            "Kế hoạch tuần": t.ke_hoach_tuan or "-",
            "Kết quả tuần": t.ket_qua_tuan or "-",
            "Vướng mắc tuần": t.vuong_mac_tuan or "-",
            "Giải quyết của CBQL/Phòng ban": t.cach_giai_quyet or "-",
            "Duyệt tuần": t.duyet_tuan or "Chưa duyệt",
            "Trạng thái": final_status_vn
        })
        
    df = pd.DataFrame(rows)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Ke_hoach_cong_viec")
        
        workbook = writer.book
        worksheet = writer.sheets["Ke_hoach_cong_viec"]
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        # Color palettes
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        level1_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
        level2_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # Sky blue 100
        level3_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Slate 50
        
        # Font configurations
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        level1_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        level2_font = Font(name="Segoe UI", size=10, bold=True, color="0369A1")
        level3_font = Font(name="Segoe UI", size=10, bold=False, color="1E293B")
        standard_font = Font(name="Segoe UI", size=10, bold=False, color="334155")
        
        # Style Header
        worksheet.row_dimensions[1].height = 28
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Style Rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(tasks) + 1), start=2):
            task_level = row[1].value  # "Cấp" is Column B (index 1)
            worksheet.row_dimensions[row_idx].height = 22
            
            # Select level styling
            if task_level == 1:
                row_fill = level1_fill
                row_font = level1_font
            elif task_level == 2:
                row_fill = level2_fill
                row_font = level2_font
            elif task_level == 3:
                row_fill = level3_fill
                row_font = level3_font
            else:
                row_fill = None
                row_font = standard_font
                
            for col_idx, cell in enumerate(row):
                if row_fill:
                    cell.fill = row_fill
                cell.font = row_font
                cell.border = thin_border
                
                # Alignments
                if col_idx in [0, 1, 2, 6, 14, 15]:  # STT, Cấp, WBS, Thời hạn, Duyệt tuần, Trạng thái
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in [8]:  # Ngân sách
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in [9]:  # Tiến độ
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
            # Number formats
            row[8].number_format = '#,##0.00" Trđ"'
            row[9].number_format = '0.0%'
            
        # Add Progress Data Bar (Green color 10B981)
        data_bar_rule = DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1.0, color="10B981", showValue=True)
        worksheet.conditional_formatting.add(f"J2:J{len(tasks)+1}", data_bar_rule)
        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            # Give column D (indented name) extra width
            if col_letter == 'D':
                worksheet.column_dimensions[col_letter].width = max(max_len + 8, 30)
            else:
                worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="VenSongVinh_WBS_KeHoach.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.get("/api/tasks/{task_id}")
def get_task(task_id: int, db: Session = Depends(database.get_db)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    # Compute rolled up deadline and cross-phase lookups properly
    all_tasks = db.query(models.Task).all()
    rolled_up_deadlines = compute_rolled_up_deadlines(all_tasks)
    rolled_date = rolled_up_deadlines.get(task.stt)
    final_deadline = format_date(rolled_date, task.thoi_han_hoan_thanh)
    
    # Get computed status
    final_status = task.trang_thai
    dot_count = task.stt.count('.')
    t_level = dot_count + 1
    if task.phase_id in (1, 3, 4):
        final_status = calculate_phase134_status(task, rolled_up_deadlines)
    elif task.phase_id == 2 and t_level == 3:
        final_status = calculate_phase2_status(task, rolled_up_deadlines)
        
    return {
        "id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "stt": task.stt,
        "phase_id": task.phase_id,
        "ten_cong_viec": task.ten_cong_viec,
        "phong_ban_thuc_hien": task.phong_ban_thuc_hien,
        "co_quan_giai_quyet": task.co_quan_giai_quyet,
        "ho_so_dau_ra": task.ho_so_dau_ra,
        "dieu_kien_ghi_nhan": task.dieu_kien_ghi_nhan,
        "thoi_han_hoan_thanh": final_deadline,
        "tien_do": task.tien_do,
        "trang_thai": final_status,
        "ke_hoach_tuan": task.ke_hoach_tuan,
        "ket_qua_tuan": task.ket_qua_tuan,
        "vuong_mac_tuan": task.vuong_mac_tuan,
        "cach_giai_quyet": task.cach_giai_quyet,
        "duyet_tuan": task.duyet_tuan,
        "weekly_reports_json": task.weekly_reports_json,
        "budget": {
            "ngan_sach_tong": task.budget.ngan_sach_tong if task.budget else 0.0,
            "is_locked": task.budget.is_locked if task.budget else False
        } if task.budget else None
    }

@app.put("/api/tasks/{task_id}/progress")
async def update_task_progress(
    task_id: int, 
    update_data: TaskProgressUpdate, 
    username: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
        
    user = None
    if username:
        user = db.query(models.User).filter(models.User.username == username).first()
        
    check_task_update_permissions(user, task, update_data, is_partial=True)
        
    # Layer 4: Phase Gate Loop check
    try:
        hooks.execute_phase_gate_loop(db, task_id, update_data.trang_thai)
    except hooks.PhaseGateException as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Update database
    task.tien_do = update_data.tien_do
    
    # Auto-calculate status unless the user is Admin
    if not user or user.role != "Admin":
        if update_data.tien_do >= 100:
            task.trang_thai = "Done"
        else:
            all_tasks = db.query(models.Task).all()
            rolled_up_deadlines = compute_rolled_up_deadlines(all_tasks)
            t_level = task.stt.count('.') + 1
            if task.phase_id in (1, 3, 4):
                task.trang_thai = calculate_phase134_status(task, rolled_up_deadlines)
            elif task.phase_id == 2 and t_level == 3:
                task.trang_thai = calculate_phase2_status(task, rolled_up_deadlines)
            else:
                task.trang_thai = "In-Progress"
    else:
        task.trang_thai = update_data.trang_thai
    if update_data.dieu_kien_ghi_nhan:
        task.dieu_kien_ghi_nhan = update_data.dieu_kien_ghi_nhan

    db.add(task)
    log_action(db, username, "Cap nhat tien do", f"Cap nhat tien do cong viec {task.stt} - {task.ten_cong_viec} thanh {task.tien_do}%. Trang thai: {task.trang_thai}.")
    db.commit()

    # Trigger real-time WebSocket broadcast
    update_msg = {
        "type": "task_update",
        "task_id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "stt": task.stt,
        "ten_cong_viec": task.ten_cong_viec,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "dieu_kien_ghi_nhan": task.dieu_kien_ghi_nhan
    }
    await manager.broadcast(update_msg)

    return {"status": "success", "task": update_msg}

@app.post("/api/spending")
async def add_spending(
    spending: SpendingCreate, 
    user_role: str = "PM", 
    db: Session = Depends(database.get_db)
):
    task = db.query(models.Task).filter(models.Task.id == spending.task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")

    # Layer 4: Hard Budget Gate check
    # If the spending is Approved, check the budget limits
    if spending.trang_thai_duyet == "Approved":
        try:
            hooks.execute_hard_budget_gate(db, spending.task_id, spending.so_tien_chi)
        except hooks.BudgetExceededException as e:
            # We notify C-level dashboard about red alert
            alert_msg = {
                "type": "red_alert",
                "message": str(e),
                "ma_ngan_sach": task.ma_ngan_sach,
                "task_id": task.id
            }
            await manager.broadcast(alert_msg)
            raise HTTPException(status_code=400, detail=str(e))

    # Save spending record
    new_spend = models.ActualSpending(
        task_id=spending.task_id,
        so_tien_chi=spending.so_tien_chi,
        ngay_chi=datetime.date.today(),
        nguoi_cap_nhat=spending.nguoi_cap_nhat,
        chung_tu_kem_theo=spending.chung_tu_kem_theo,
        trang_thai_duyet=spending.trang_thai_duyet
    )
    db.add(new_spend)
    log_action(db, spending.nguoi_cap_nhat, "De xuat giai ngan", f"Giai ngan thuc te {spending.so_tien_chi} Trd cho cong viec {task.stt} - {task.ten_cong_viec}.")
    db.commit()

    # Broadcast spending update
    spending_msg = {
        "type": "spending_update",
        "task_id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "so_tien_chi": spending.so_tien_chi,
        "nguoi_cap_nhat": spending.nguoi_cap_nhat,
        "trang_thai_duyet": spending.trang_thai_duyet
    }
    await manager.broadcast(spending_msg)

    return {"status": "success", "spending_id": new_spend.id}

@app.get("/api/spending")
def get_spending(db: Session = Depends(database.get_db)):
    spendings = db.query(models.ActualSpending).all()
    result = []
    for s in spendings:
        task = db.query(models.Task).filter(models.Task.id == s.task_id).first()
        result.append({
            "id": s.id,
            "task_id": s.task_id,
            "ma_ngan_sach": task.ma_ngan_sach if task else "N/A",
            "ten_cong_viec": task.ten_cong_viec if task else "N/A",
            "so_tien_chi": s.so_tien_chi,
            "ngay_chi": s.ngay_chi.isoformat(),
            "nguoi_cap_nhat": s.nguoi_cap_nhat,
            "chung_tu_kem_theo": s.chung_tu_kem_theo,
            "trang_thai_duyet": s.trang_thai_duyet
        })
    return result

@app.get("/api/stats")
def get_dashboard_stats(db: Session = Depends(database.get_db)):
    project = db.query(models.Project).filter(models.Project.id == 1).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Total budget
    total_budget = project.tong_ngan_sach

    # Total actual spending approved
    total_spent = db.query(models.ActualSpending).filter(
        models.ActualSpending.trang_thai_duyet == "Approved"
    ).all()
    total_spent_val = sum(s.so_tien_chi for s in total_spent)

    # Average progress of Level 1 tasks
    lvl_1_tasks = db.query(models.Task).all()
    # Filter Level 1
    lvl_1_tasks = [t for t in lvl_1_tasks if len(t.stt.split('.')) == 1]
    avg_progress = sum(t.tien_do for t in lvl_1_tasks) / len(lvl_1_tasks) if lvl_1_tasks else 0.0

    # Count tasks by status
    all_tasks = db.query(models.Task).all()
    todo_count = sum(1 for t in all_tasks if t.trang_thai == "Todo")
    inprogress_count = sum(1 for t in all_tasks if t.trang_thai == "In-Progress")
    done_count = sum(1 for t in all_tasks if t.trang_thai == "Done")
    delayed_count = sum(1 for t in all_tasks if t.trang_thai == "Delayed")

    # Locked budgets count
    locked_count = db.query(models.Budget).filter(models.Budget.is_locked == True).count()

    return {
        "tong_ngan_sach": total_budget,
        "tong_thuc_chi": total_spent_val,
        "con_lai": total_budget - total_spent_val,
        "tien_do_trung_binh": avg_progress,
        "trang_thai_tasks": {
            "Todo": todo_count,
            "In-Progress": inprogress_count,
            "Done": done_count,
            "Delayed": delayed_count
        },
        "so_wbs_bi_khoa": locked_count
    }

@app.get("/api/s-curve")
def get_s_curve_data(db: Session = Depends(database.get_db)):
    """
    Tạo dữ liệu đường cong S-Curve kế hoạch vs thực tế chi tiêu.
    Các mốc thời gian: 
    - 01/2025, 02/2025, 03/2025, 04/2025
    - Quý 1/2026 (biểu diễn mốc 03/2026)
    - 05/2026, 06/2026
    """
    budgets = db.query(models.Budget).all()
    
    # Kế hoạch chi tiêu cho từng mốc (tính tổng của toàn bộ các task)
    plan_t1_2025 = sum(b.thang_01_2025 for b in budgets)
    plan_t2_2025 = sum(b.thang_02_2025 for b in budgets)
    plan_t3_2025 = sum(b.thang_03_2025 for b in budgets)
    plan_t4_2025 = sum(b.thang_04_2025 for b in budgets)
    plan_q1_2026 = sum(b.quy_1_2026 for b in budgets)
    plan_t5_2026 = sum(b.thang_05_2026 for b in budgets)
    plan_t6_2026 = sum(b.thang_06_2026 for b in budgets)

    # Lũy kế kế hoạch chi tiêu
    cum_plan = []
    
    # 01/2025
    cum_plan.append({"period": "Tháng 01/2025", "value": plan_t1_2025})
    # 02/2025
    cum_plan.append({"period": "Tháng 02/2025", "value": plan_t1_2025 + plan_t2_2025})
    # 03/2025
    cum_plan.append({"period": "Tháng 03/2025", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025})
    # 04/2025
    cum_plan.append({"period": "Tháng 04/2025", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025})
    # Quý 1/2026
    cum_plan.append({"period": "Quý 1/2026", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025 + plan_q1_2026})
    # 05/2026
    cum_plan.append({"period": "Tháng 05/2026", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025 + plan_q1_2026 + plan_t5_2026})
    # 06/2026
    cum_plan.append({"period": "Tháng 06/2026", "value": plan_t1_2025 + plan_t2_2025 + plan_t3_2025 + plan_t4_2025 + plan_q1_2026 + plan_t5_2026 + plan_t6_2026})

    # Lũy kế thực chi theo mốc thời gian thực tế nhập trong actual_spending
    spendings = db.query(models.ActualSpending).filter(
        models.ActualSpending.trang_thai_duyet == "Approved"
    ).all()

    # Phân bổ thực chi vào các mốc thời gian tương ứng
    actual_t1_2025 = 0.0
    actual_t2_2025 = 0.0
    actual_t3_2025 = 0.0
    actual_t4_2025 = 0.0
    actual_q1_2026 = 0.0
    actual_t5_2026 = 0.0
    actual_t6_2026 = 0.0

    for s in spendings:
        date = s.ngay_chi
        # Phân chia thời gian giải ngân
        if date.year == 2025:
            if date.month == 1:
                actual_t1_2025 += s.so_tien_chi
            elif date.month == 2:
                actual_t2_2025 += s.so_tien_chi
            elif date.month == 3:
                actual_t3_2025 += s.so_tien_chi
            elif date.month == 4:
                actual_t4_2025 += s.so_tien_chi
            else:
                # Gán tạm vào tháng 4 nếu là các tháng khác của 2025
                actual_t4_2025 += s.so_tien_chi
        elif date.year == 2026:
            if date.month in [1, 2, 3]:
                actual_q1_2026 += s.so_tien_chi
            elif date.month == 5:
                actual_t5_2026 += s.so_tien_chi
            elif date.month == 6:
                actual_t6_2026 += s.so_tien_chi
            else:
                actual_t6_2026 += s.so_tien_chi
        else:
            # Năm 2027+
            actual_t6_2026 += s.so_tien_chi

    cum_actual = []
    # 01/2025
    cum_actual.append({"period": "Tháng 01/2025", "value": actual_t1_2025})
    # 02/2025
    cum_actual.append({"period": "Tháng 02/2025", "value": actual_t1_2025 + actual_t2_2025})
    # 03/2025
    cum_actual.append({"period": "Tháng 03/2025", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025})
    # 04/2025
    cum_actual.append({"period": "Tháng 04/2025", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025})
    # Quý 1/2026
    cum_actual.append({"period": "Quý 1/2026", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025 + actual_q1_2026})
    # 05/2026
    cum_actual.append({"period": "Tháng 05/2026", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025 + actual_q1_2026 + actual_t5_2026})
    # 06/2026
    cum_actual.append({"period": "Tháng 06/2026", "value": actual_t1_2025 + actual_t2_2025 + actual_t3_2025 + actual_t4_2025 + actual_q1_2026 + actual_t5_2026 + actual_t6_2026})

    return {
        "plan": cum_plan,
        "actual": cum_actual
    }

@app.post("/api/ai/parse")
async def parse_and_update_task(payload: AIUpdateInput, db: Session = Depends(database.get_db)):
    """
    AI Integration Endpoint
    Nhận văn bản thô -> Gemini bóc tách -> Tự động cập nhật Task thành Done và điền điều kiện ghi nhận
    """
    result = ai_agent.parse_natural_language_update(payload.text, payload.api_key)
    
    ma_ngan_sach = result.get("ma_ngan_sach", "")
    trang_thai = result.get("trang_thai", "Done")
    tien_do = result.get("tien_do", 100.0)
    dieu_kien = result.get("dieu_kien_ghi_nhan", "")

    if not ma_ngan_sach:
        raise HTTPException(status_code=400, detail="AI không thể phát hiện mã ngân sách/STT phù hợp trong văn bản.")

    # Tìm task phù hợp trong DB (kiểm tra cả ma_ngan_sach và stt)
    task = db.query(models.Task).filter(
        (models.Task.ma_ngan_sach == ma_ngan_sach) |
        (models.Task.stt == ma_ngan_sach)
    ).first()

    if not task:
        # Thử tìm theo dạng chứa phần cuối của STT
        task = db.query(models.Task).filter(models.Task.stt.like(f"%{ma_ngan_sach}%")).first()

    if not task:
        return {
            "status": "partial_success",
            "message": f"AI trích xuất thành công nhưng không tìm thấy công việc tương ứng cho mã '{ma_ngan_sach}'.",
            "extracted_data": result
        }

    # Chạy chốt chặn Phase Gate Loop trước khi cập nhật
    try:
        hooks.execute_phase_gate_loop(db, task.id, trang_thai)
    except hooks.PhaseGateException as e:
        raise HTTPException(status_code=400, detail=f"Lỗi Logic Chốt chặn: {str(e)}")

    # Tiến hành cập nhật
    task.tien_do = tien_do
    task.trang_thai = trang_thai
    task.dieu_kien_ghi_nhan = dieu_kien
    
    db.add(task)
    db.commit()

    # Broadcast WebSocket update
    update_msg = {
        "type": "task_update",
        "task_id": task.id,
        "ma_ngan_sach": task.ma_ngan_sach,
        "stt": task.stt,
        "ten_cong_viec": task.ten_cong_viec,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "dieu_kien_ghi_nhan": task.dieu_kien_ghi_nhan
    }
    await manager.broadcast(update_msg)

    return {
        "status": "success",
        "message": f"Đã tự động cập nhật công việc {task.stt} thành công qua {result['method']}.",
        "task": update_msg
    }

@app.get("/api/ai/risk")
def get_ai_risk_assessment(api_key: Optional[str] = None, db: Session = Depends(database.get_db)):
    project = db.query(models.Project).filter(models.Project.id == 1).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    total_spent = db.query(models.ActualSpending).filter(
        models.ActualSpending.trang_thai_duyet == "Approved"
    ).all()
    total_spent_val = sum(s.so_tien_chi for s in total_spent)

    # Lấy thông tin 10 khoản chi gần nhất để AI phân tích
    spending_details = []
    for s in total_spent[-10:]:
        task = db.query(models.Task).filter(models.Task.id == s.task_id).first()
        spending_details.append({
            "ma_ngan_sach": task.ma_ngan_sach if task else "N/A",
            "ten_cong_viec": task.ten_cong_viec if task else "N/A",
            "so_tien_chi": s.so_tien_chi,
            "nguoi_chi": s.nguoi_cap_nhat
        })

    risk_report = ai_agent.evaluate_financial_risk(
        project.tong_ngan_sach,
        total_spent_val,
        spending_details,
        api_key
    )
    
    return {"risk_report": risk_report}

class AIChatRequest(BaseModel):
    question: str
    context: str
    api_key: Optional[str] = None

@app.post("/api/ai/chat")
async def chat_with_gemini(request: AIChatRequest):
    api_key = request.api_key or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        raise HTTPException(status_code=400, detail="Gemini API Key is missing. Please configure it in the sidebar.")
    
    prompt = f"""
Bạn là Chuyên gia Kiến trúc Hệ thống ERP và Phân tích Dữ liệu AI của Dự án Bất động sản Ven Sông Vinh.
Dưới đây là dữ liệu thực tế thời gian thực của dự án:
{request.context}

Câu hỏi của người dùng:
"{request.question}"

Hãy phân tích dữ liệu và trả lời câu hỏi một cách ngắn gọn, súc tích, chuyên nghiệp bằng Tiếng Việt.
"""
    try:
        response_text = ai_agent.generate_generic_text(prompt, api_key=api_key)
        return {"response": response_text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class TaskCreate(BaseModel):
    parent_stt: str
    ten_cong_viec: str
    phong_ban_thuc_hien: Optional[str] = "BQLDA"
    co_quan_giai_quyet: Optional[str] = "-"
    ho_so_dau_ra: Optional[str] = "-"
    dieu_kien_ghi_nhan: Optional[str] = "-"
    thoi_han_hoan_thanh: Optional[str] = "Tháng 06/2026"
    ngan_sach: float = 0.0

def recalculate_budgets(db: Session):
    remaining_tasks = db.query(models.Task).all()
    task_by_stt = {t.stt: t for t in remaining_tasks}

    all_stts = {t.stt for t in remaining_tasks}
    parent_tasks = []
    
    for t in remaining_tasks:
        is_parent = any(other.startswith(f"{t.stt}.") for other in all_stts)
        if is_parent:
            parent_tasks.append(t)
            
    for t in parent_tasks:
        if t.budget:
            b = t.budget
            b.ngan_sach_tong = 0.0
            b.kh_2026 = 0.0
            b.quy_1_2026 = 0.0
            b.quy_2_2026 = 0.0
            b.thang_01_2025 = 0.0
            b.thang_02_2025 = 0.0
            b.thang_03_2025 = 0.0
            b.thang_04_2025 = 0.0
            b.thang_05_2026 = 0.0
            b.thang_06_2026 = 0.0
            db.add(b)
    db.flush()

    sorted_tasks = sorted(remaining_tasks, key=lambda t: len(t.stt.split('.')), reverse=True)

    for task in sorted_tasks:
        stt_parts = task.stt.split('.')
        if len(stt_parts) <= 1:
            continue
        
        parent_stt = ".".join(stt_parts[:-1])
        parent_task = task_by_stt.get(parent_stt)

        if parent_task:
            child_budget = task.budget
            parent_budget = parent_task.budget
            
            if child_budget and parent_budget:
                parent_budget.ngan_sach_tong += child_budget.ngan_sach_tong
                parent_budget.kh_2026 += child_budget.kh_2026
                parent_budget.quy_1_2026 += child_budget.quy_1_2026
                parent_budget.quy_2_2026 += child_budget.quy_2_2026
                parent_budget.thang_01_2025 += child_budget.thang_01_2025
                parent_budget.thang_02_2025 += child_budget.thang_02_2025
                parent_budget.thang_03_2025 += child_budget.thang_03_2025
                parent_budget.thang_04_2025 += child_budget.thang_04_2025
                parent_budget.thang_05_2026 += child_budget.thang_05_2026
                parent_budget.thang_06_2026 += child_budget.thang_06_2026
                db.add(parent_budget)
    db.flush()

    level_1_tasks = [t for t in remaining_tasks if len(t.stt.split('.')) == 1]
    project_total = sum(t.budget.ngan_sach_tong for t in level_1_tasks if t.budget)

    project_obj = db.query(models.Project).filter(models.Project.id == 1).first()
    if project_obj:
        project_obj.tong_ngan_sach = project_total
        db.add(project_obj)
    db.flush()

def parse_phase_from_stt(stt_val):
    if not stt_val:
        return 1
    stt_str = str(stt_val).strip()
    parts = stt_str.split('.')
    if not parts or not parts[0].isdigit():
        return 1
    try:
        prefix = int(parts[0])
        if 1 <= prefix <= 15:
            return 1
        elif 16 <= prefix <= 18:
            return 2
        elif 19 <= prefix <= 26:
            return 3
        elif 27 <= prefix <= 28:
            return 4
        else:
            return 1
    except ValueError:
        return 1

@app.post("/api/tasks/import-excel-ai")
async def import_excel_ai(
    file: UploadFile = File(...),
    username: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    try:
        # Read the file contents
        file_bytes = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        
        # Get first 15 rows for AI analysis
        rows_data = []
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            serialized_row = []
            for cell in row:
                if cell is None:
                    serialized_row.append("")
                elif isinstance(cell, datetime.datetime) or isinstance(cell, datetime.date):
                    serialized_row.append(cell.isoformat())
                else:
                    serialized_row.append(str(cell))
            rows_data.append(serialized_row)
            if len(rows_data) >= 15:
                break
                
        # Get Gemini API key if configured
        api_key = os.environ.get("GEMINI_API_KEY")
        
        # Call AI bóc tách mapping
        mapping = ai_agent.extract_excel_mapping_with_ai(rows_data, api_key)
        print(f"AI Column Mapping determined: {mapping}")
        
        # Column mappings (0-based indices)
        ma_ngan_sach_idx = mapping.get("ma_ngan_sach_idx")
        stt_idx = mapping.get("stt_idx")
        ten_cong_viec_idx = mapping.get("ten_cong_viec_idx")
        phong_ban_idx = mapping.get("phong_ban_idx")
        co_quan_idx = mapping.get("co_quan_idx")
        ho_so_dau_ra_idx = mapping.get("ho_so_dau_ra_idx")
        dieu_kien_ghi_nhan_idx = mapping.get("dieu_kien_ghi_nhan_idx")
        thoi_han_hoan_thanh_idx = mapping.get("thoi_han_hoan_thanh_idx")
        tien_do_idx = mapping.get("tien_do_idx")
        trang_thai_idx = mapping.get("trang_thai_idx")
        ngan_sach_idx = mapping.get("ngan_sach_idx")
        
        # If crucial columns (WBS/STT or Name) are missing, we cannot parse
        if ma_ngan_sach_idx is None or ten_cong_viec_idx is None:
            raise HTTPException(
                status_code=400,
                detail="Không thể tự động bóc tách cấu trúc cột. Vui lòng kiểm tra lại file Excel có chứa cột Mã ngân sách/WBS và Tên công việc."
            )
            
        # Determine the start row (usually we skip headers)
        start_row = 1
        for idx, row in enumerate(rows_data[:5]):
            row_str = " ".join([str(x).lower() for x in row if x])
            if "stt" in row_str or "wbs" in row_str or "mã ngân sách" in row_str or "tên công việc" in row_str:
                start_row = idx + 2
                break
                
        # Parse rows
        excel_rows = list(sheet.iter_rows(values_only=True))
        added_count = 0
        updated_count = 0
        
        for idx, row in enumerate(excel_rows):
            r = idx + 1
            if r < start_row:
                continue
                
            wbs_val = str(row[ma_ngan_sach_idx]).strip() if ma_ngan_sach_idx is not None and row[ma_ngan_sach_idx] is not None else ""
            stt_val = str(row[stt_idx]).strip() if stt_idx is not None and row[stt_idx] is not None else f"ROW{r}"
            name_val = str(row[ten_cong_viec_idx]).strip() if ten_cong_viec_idx is not None and row[ten_cong_viec_idx] is not None else ""
            
            if not name_val or name_val.upper() == "TỔNG DỰ ÁN":
                continue
                
            if not wbs_val:
                wbs_val = f"TD.BĐS.GEN.{stt_val}"
                
            phong_ban_val = str(row[phong_ban_idx]).strip() if phong_ban_idx is not None and row[phong_ban_idx] is not None else "PTDA"
            co_quan_val = str(row[co_quan_idx]).strip() if co_quan_idx is not None and row[co_quan_idx] is not None else "-"
            ho_so_val = str(row[ho_so_dau_ra_idx]).strip() if ho_so_dau_ra_idx is not None and row[ho_so_dau_ra_idx] is not None else "-"
            dieu_kien_val = str(row[dieu_kien_ghi_nhan_idx]).strip() if dieu_kien_ghi_nhan_idx is not None and row[dieu_kien_ghi_nhan_idx] is not None else "-"
            deadline_val = str(row[thoi_han_hoan_thanh_idx]).strip() if thoi_han_hoan_thanh_idx is not None and row[thoi_han_hoan_thanh_idx] is not None else ""
            
            tien_do_val = 0.0
            if tien_do_idx is not None and row[tien_do_idx] is not None:
                try:
                    val = str(row[tien_do_idx]).replace('%', '').strip()
                    tien_do_val = float(val)
                    if tien_do_val > 1.0 and tien_do_val <= 100.0:
                        pass
                    elif tien_do_val >= 0.0 and tien_do_val <= 1.0:
                        tien_do_val = tien_do_val * 100.0
                except ValueError:
                    pass
                    
            trang_thai_val = "Todo"
            if trang_thai_idx is not None and row[trang_thai_idx] is not None:
                status_raw = str(row[trang_thai_idx]).lower().strip()
                if "done" in status_raw or "hoàn thành" in status_raw or "xong" in status_raw:
                    trang_thai_val = "Done"
                elif "progress" in status_raw or "đang" in status_raw or "triển khai" in status_raw:
                    trang_thai_val = "In-Progress"
                elif "delayed" in status_raw or "trễ" in status_raw or "chậm" in status_raw:
                    trang_thai_val = "Delayed"
            else:
                if tien_do_val >= 100.0:
                    trang_thai_val = "Done"
                elif tien_do_val > 0.0:
                    trang_thai_val = "In-Progress"
                    
            budget_val = 0.0
            if ngan_sach_idx is not None and row[ngan_sach_idx] is not None:
                try:
                    val = str(row[ngan_sach_idx]).replace(',', '').strip()
                    budget_val = float(val.replace('.', '')) if '.' in val and ',' in val else float(val)
                except ValueError:
                    pass
            
            phase_id = parse_phase_from_stt(stt_val)
            
            # Upsert logic
            existing_task = db.query(models.Task).filter(models.Task.ma_ngan_sach == wbs_val).first()
            if existing_task:
                existing_task.stt = stt_val
                existing_task.ten_cong_viec = name_val
                existing_task.phong_ban_thuc_hien = phong_ban_val
                existing_task.co_quan_giai_quyet = co_quan_val
                existing_task.ho_so_dau_ra = ho_so_val
                existing_task.dieu_kien_ghi_nhan = dieu_kien_val
                existing_task.thoi_han_hoan_thanh = deadline_val
                existing_task.tien_do = tien_do_val
                existing_task.trang_thai = trang_thai_val
                existing_task.phase_id = phase_id
                
                if existing_task.budget:
                    existing_task.budget.ngan_sach_tong = budget_val
                else:
                    new_budget = models.Budget(
                        task_id=existing_task.id,
                        ngan_sach_tong=budget_val,
                        kh_2026=budget_val * 0.40,
                        quy_1_2026=budget_val * 0.15,
                        quy_2_2026=budget_val * 0.15,
                        thang_01_2025=budget_val * 0.05,
                        thang_02_2025=budget_val * 0.05,
                        thang_03_2025=budget_val * 0.05,
                        thang_04_2025=budget_val * 0.05,
                        thang_05_2026=budget_val * 0.05,
                        thang_06_2026=budget_val * 0.05
                    )
                    db.add(new_budget)
                updated_count += 1
            else:
                new_task = models.Task(
                    project_id=1,
                    ma_ngan_sach=wbs_val,
                    stt=stt_val,
                    phase_id=phase_id,
                    ten_cong_viec=name_val,
                    phong_ban_thuc_hien=phong_ban_val,
                    co_quan_giai_quyet=co_quan_val,
                    ho_so_dau_ra=ho_so_val,
                    dieu_kien_ghi_nhan=dieu_kien_val,
                    thoi_han_hoan_thanh=deadline_val,
                    tien_do=tien_do_val,
                    trang_thai=trang_thai_val,
                    ngay_khoi_tao=datetime.date.today().isoformat(),
                    weekly_reports_json="[]"
                )
                db.add(new_task)
                db.flush()
                
                new_budget = models.Budget(
                    task_id=new_task.id,
                    ngan_sach_tong=budget_val,
                    kh_2026=budget_val * 0.40,
                    quy_1_2026=budget_val * 0.15,
                    quy_2_2026=budget_val * 0.15,
                    thang_01_2025=budget_val * 0.05,
                    thang_02_2025=budget_val * 0.05,
                    thang_03_2025=budget_val * 0.05,
                    thang_04_2025=budget_val * 0.05,
                    thang_05_2026=budget_val * 0.05,
                    thang_06_2026=budget_val * 0.05
                )
                db.add(new_budget)
                added_count += 1
                
        recalculate_budgets(db)
        log_action(db, username or "Hệ thống AI", "Bóc tách Excel", f"Đã nhập Excel thành công: Thêm mới {added_count} dòng, Cập nhật {updated_count} dòng.")
        db.commit()
        
        await manager.broadcast({"type": "sync_refresh"})
        
        return {
            "status": "success",
            "message": f"Bóc tách Excel thành công! Đã thêm mới {added_count} công việc, Cập nhật {updated_count} công việc.",
            "mapping": mapping
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Lỗi bóc tách file Excel: {str(e)}")

@app.post("/api/tasks")
async def create_task(
    request: TaskCreate, 
    username: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    parent = db.query(models.Task).filter(models.Task.stt == request.parent_stt).first()
    if not parent:
        raise HTTPException(status_code=404, detail=f"Không tìm thấy công việc cha có STT {request.parent_stt}")
        
    parent_parts = parent.stt.split('.')
    parent_level = len(parent_parts)
    child_level = parent_level + 1
    
    user = None
    if username:
        user = db.query(models.User).filter(models.User.username == username).first()
        
    # Rule 1: Only Admin can add Level 1 & 2 tasks (child_level <= 2)
    if child_level <= 2:
        if not user or user.role != "Admin":
            raise HTTPException(
                status_code=403,
                detail="Chỉ có Admin mới được phép thêm công việc cấp 1, cấp 2."
            )
            
    # Rule 2 & 3: Level >= 3 tasks
    if child_level >= 3:
        if not user:
            raise HTTPException(status_code=403, detail="Yêu cầu đăng nhập để thêm công việc con.")
            
        user_dept = str(user.phong_ban).upper().strip()
        parent_dept = str(parent.phong_ban_thuc_hien).upper().strip()
        
        # Non-admins must belong to the department of the parent task
        if user.role != "Admin" and user_dept != parent_dept and user_dept != "ALL":
            raise HTTPException(
                status_code=403,
                detail=f"Quyền hạn bị từ chối: Phòng ban của bạn ({user.phong_ban}) không trùng khớp với phòng thực hiện công việc cha ({parent.phong_ban_thuc_hien})."
            )
            
        # Level 3 tasks: only Admin or Trưởng phòng (TruongPhong/CBQL) can create
        if child_level == 3 and user.role != "Admin":
            if user.role not in ("TruongPhong", "CBQL", "PM"):
                raise HTTPException(
                    status_code=403,
                    detail="Chỉ có Admin hoặc Trưởng phòng mới được phép thêm công việc cấp 3."
                )
            
        # Level 4 and 5 tasks: Admin, Trưởng phòng, or NhanVien of the same department can create (all are allowed if department matches)
        
    siblings = db.query(models.Task).filter(models.Task.stt.like(f"{request.parent_stt}.%")).all()
    sibling_indices = []
    for s in siblings:
        parts = s.stt.split('.')
        if len(parts) == parent_level + 1 and parts[-1].isdigit():
            sibling_indices.append(int(parts[-1]))
            
    next_idx = max(sibling_indices) + 1 if sibling_indices else 1
    new_stt = f"{request.parent_stt}.{next_idx}"
    new_wbs = f"{parent.ma_ngan_sach}.{next_idx}"
    
    creator_name = user.ho_ten if user else (username or "Hệ thống Admin")
    
    # Lookup CBQL (Trưởng phòng) of the department
    dept_val = request.phong_ban_thuc_hien.strip()
    dept_cbql = db.query(models.User).filter(
        models.User.phong_ban == dept_val,
        models.User.role == "TruongPhong"
    ).first()
    cbql_name = dept_cbql.ho_ten if dept_cbql else "-"
    
    new_task = models.Task(
        project_id=1,
        ma_ngan_sach=new_wbs,
        stt=new_stt,
        phase_id=parent.phase_id,
        ten_cong_viec=request.ten_cong_viec.strip(),
        phong_ban_thuc_hien=dept_val,
        co_quan_giai_quyet=request.co_quan_giai_quyet.strip(),
        ho_so_dau_ra=request.ho_so_dau_ra.strip(),
        dieu_kien_ghi_nhan=request.dieu_kien_ghi_nhan.strip(),
        thoi_han_hoan_thanh=request.thoi_han_hoan_thanh.strip(),
        tien_do=0.0,
        trang_thai="Todo",
        nguoi_phu_trach=creator_name,
        nguoi_bao_cao=creator_name,
        nguoi_duyet=cbql_name,
        ngay_khoi_tao=datetime.date.today().isoformat()
    )
    db.add(new_task)
    db.flush()
    
    new_budget = models.Budget(
        task_id=new_task.id,
        ngan_sach_tong=request.ngan_sach,
        kh_2026=request.ngan_sach * 0.40,
        quy_1_2026=request.ngan_sach * 0.15,
        quy_2_2026=request.ngan_sach * 0.15,
        thang_01_2025=request.ngan_sach * 0.05,
        thang_02_2025=request.ngan_sach * 0.05,
        thang_03_2025=request.ngan_sach * 0.05,
        thang_04_2025=request.ngan_sach * 0.05,
        thang_05_2026=request.ngan_sach * 0.05,
        thang_06_2026=request.ngan_sach * 0.05
    )
    db.add(new_budget)
    db.flush()
    
    recalculate_budgets(db)
    log_action(db, username, "Thêm mới công việc con", f"Thêm công việc con {new_stt} - {new_task.ten_cong_viec} dưới WBS cha {request.parent_stt}")
    db.commit()
    
    update_msg = {
        "type": "task_create",
        "task_id": new_task.id,
        "stt": new_task.stt,
        "ma_ngan_sach": new_task.ma_ngan_sach,
        "ten_cong_viec": new_task.ten_cong_viec,
        "phong_ban_thuc_hien": new_task.phong_ban_thuc_hien,
        "ho_so_dau_ra": new_task.ho_so_dau_ra,
        "thoi_han_hoan_thanh": new_task.thoi_han_hoan_thanh,
        "tien_do": new_task.tien_do,
        "trang_thai": new_task.trang_thai,
        "budget": {
            "ngan_sach_tong": new_budget.ngan_sach_tong,
            "is_locked": False
        }
    }
    await manager.broadcast(update_msg)
    
    return {
        "status": "success",
        "message": f"Đã thêm mới thành công công việc Cấp 3: {new_stt}",
        "task": update_msg
    }

class TaskUpdate(BaseModel):
    ma_ngan_sach: str
    ten_cong_viec: str
    phong_ban_thuc_hien: str
    co_quan_giai_quyet: str
    ho_so_dau_ra: str
    dieu_kien_ghi_nhan: str
    thoi_han_hoan_thanh: str
    tien_do: float
    trang_thai: str
    ngan_sach: float
    ke_hoach_tuan: Optional[str] = ""
    ket_qua_tuan: Optional[str] = ""
    vuong_mac_tuan: Optional[str] = ""
    cach_giai_quyet: Optional[str] = ""
    duyet_tuan: Optional[str] = "Chưa duyệt"
    
    # New Operational and Financial fields
    ngay_khoi_tao: Optional[str] = ""
    cong_trinh: Optional[str] = ""
    doi_tac: Optional[str] = ""
    so_dien_thoai: Optional[str] = ""
    ngay_bat_dau: Optional[str] = ""
    ngay_ket_thuc: Optional[str] = ""
    gia_han_den_ngay: Optional[str] = ""
    thoi_gian_bao_hanh: Optional[str] = ""
    mo_ta: Optional[str] = ""
    dieu_khoan: Optional[str] = ""
    
    nguoi_phu_trach: Optional[str] = ""
    nguoi_bao_cao: Optional[str] = ""
    nguoi_duyet: Optional[str] = ""
    
    gia_tri_quyet_toan: Optional[float] = 0.0
    da_nghiem_thu: Optional[float] = 0.0
    da_thanh_toan: Optional[float] = 0.0
    tam_ung: Optional[float] = 0.0
    da_thu_hoi_tam_ung: Optional[float] = 0.0
    weekly_reports_json: Optional[str] = "[]"

@app.put("/api/tasks/{task_id}")
async def update_task_details(
    task_id: int, 
    request: TaskUpdate, 
    username: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")
        
    user = None
    if username:
        user = db.query(models.User).filter(models.User.username == username).first()
        
    check_task_update_permissions(user, task, request, is_partial=False)
    
    # Capture old values
    old_wbs = task.ma_ngan_sach or ""
    old_name = task.ten_cong_viec or ""
    old_dept = task.phong_ban_thuc_hien or ""
    old_deliverables = task.ho_so_dau_ra or ""
    old_cond = task.dieu_kien_ghi_nhan or ""
    old_deadline = task.thoi_han_hoan_thanh or ""
    old_progress = task.tien_do or 0.0
    old_status = task.trang_thai or "Todo"
    old_budget = task.budget.ngan_sach_tong if task.budget else 0.0
    
    old_plan = task.ke_hoach_tuan or ""
    old_result = task.ket_qua_tuan or ""
    old_issues = task.vuong_mac_tuan or ""
    
    old_solution = task.cach_giai_quyet or ""
    old_approval = task.duyet_tuan or "Chưa duyệt"
        
    if request.ma_ngan_sach != task.ma_ngan_sach:
        existing = db.query(models.Task).filter(models.Task.ma_ngan_sach == request.ma_ngan_sach).first()
        if existing:
            raise HTTPException(status_code=400, detail="Mã Ngân Sách (WBS) này đã tồn tại.")
            
    task.ma_ngan_sach = request.ma_ngan_sach.strip()
    task.ten_cong_viec = request.ten_cong_viec.strip()
    task.phong_ban_thuc_hien = request.phong_ban_thuc_hien.strip()
    task.co_quan_giai_quyet = request.co_quan_giai_quyet.strip()
    task.ho_so_dau_ra = request.ho_so_dau_ra.strip()
    task.dieu_kien_ghi_nhan = request.dieu_kien_ghi_nhan.strip()
    task.thoi_han_hoan_thanh = request.thoi_han_hoan_thanh.strip()
    task.tien_do = request.tien_do
    
    # Auto-calculate status unless the user is Admin
    if not user or user.role != "Admin":
        if request.tien_do >= 100:
            task.trang_thai = "Done"
        else:
            all_tasks = db.query(models.Task).all()
            rolled_up_deadlines = compute_rolled_up_deadlines(all_tasks)
            t_level = task.stt.count('.') + 1
            if task.phase_id in (1, 3, 4):
                task.trang_thai = calculate_phase134_status(task, rolled_up_deadlines)
            elif task.phase_id == 2 and t_level == 3:
                task.trang_thai = calculate_phase2_status(task, rolled_up_deadlines)
            else:
                task.trang_thai = "In-Progress"
    else:
        task.trang_thai = request.trang_thai
    task.ke_hoach_tuan = request.ke_hoach_tuan.strip() if request.ke_hoach_tuan else ""
    task.ket_qua_tuan = request.ket_qua_tuan.strip() if request.ket_qua_tuan else ""
    task.vuong_mac_tuan = request.vuong_mac_tuan.strip() if request.vuong_mac_tuan else ""
    task.cach_giai_quyet = request.cach_giai_quyet.strip() if request.cach_giai_quyet else ""
    task.duyet_tuan = request.duyet_tuan.strip() if request.duyet_tuan else "Chưa duyệt"
    
    # Map new fields
    task.ngay_khoi_tao = request.ngay_khoi_tao.strip() if request.ngay_khoi_tao else ""
    task.cong_trinh = request.cong_trinh.strip() if request.cong_trinh else ""
    task.doi_tac = request.doi_tac.strip() if request.doi_tac else ""
    task.so_dien_thoai = request.so_dien_thoai.strip() if request.so_dien_thoai else ""
    task.ngay_bat_dau = request.ngay_bat_dau.strip() if request.ngay_bat_dau else ""
    task.ngay_ket_thuc = request.ngay_ket_thuc.strip() if request.ngay_ket_thuc else ""
    task.gia_han_den_ngay = request.gia_han_den_ngay.strip() if request.gia_han_den_ngay else ""
    task.thoi_gian_bao_hanh = request.thoi_gian_bao_hanh.strip() if request.thoi_gian_bao_hanh else ""
    task.mo_ta = request.mo_ta.strip() if request.mo_ta else ""
    task.dieu_khoan = request.dieu_khoan.strip() if request.dieu_khoan else ""
    
    task.nguoi_phu_trach = request.nguoi_phu_trach.strip() if request.nguoi_phu_trach else ""
    task.nguoi_bao_cao = request.nguoi_bao_cao.strip() if request.nguoi_bao_cao else ""
    task.nguoi_duyet = request.nguoi_duyet.strip() if request.nguoi_duyet else ""
    
    task.gia_tri_quyet_toan = request.gia_tri_quyet_toan or 0.0
    task.da_nghiem_thu = request.da_nghiem_thu or 0.0
    task.da_thanh_toan = request.da_thanh_toan or 0.0
    task.tam_ung = request.tam_ung or 0.0
    task.da_thu_hoi_tam_ung = request.da_thu_hoi_tam_ung or 0.0
    task.weekly_reports_json = request.weekly_reports_json.strip() if request.weekly_reports_json else "[]"
    
    if task.budget:
        task.budget.ngan_sach_tong = request.ngan_sach
        task.budget.kh_2026 = request.ngan_sach * 0.40
        task.budget.quy_1_2026 = request.ngan_sach * 0.15
        task.budget.quy_2_2026 = request.ngan_sach * 0.15
        task.budget.thang_01_2025 = request.ngan_sach * 0.05
        task.budget.thang_02_2025 = request.ngan_sach * 0.05
        task.budget.thang_03_2025 = request.ngan_sach * 0.05
        task.budget.thang_04_2025 = request.ngan_sach * 0.05
        task.budget.thang_05_2026 = request.ngan_sach * 0.05
        task.budget.thang_06_2026 = request.ngan_sach * 0.05
    db.add(task)
    db.flush()
    
    is_type_a_changed = (
        old_wbs != task.ma_ngan_sach or
        old_name != task.ten_cong_viec or
        old_dept != task.phong_ban_thuc_hien or
        old_deliverables != task.ho_so_dau_ra or
        old_cond != task.dieu_kien_ghi_nhan or
        old_deadline != task.thoi_han_hoan_thanh or
        old_progress != task.tien_do or
        old_status != task.trang_thai or
        (task.budget and old_budget != task.budget.ngan_sach_tong)
    )
    
    is_type_b_changed = (
        old_plan != (task.ke_hoach_tuan or "") or
        old_result != (task.ket_qua_tuan or "") or
        old_issues != (task.vuong_mac_tuan or "")
    )
    
    is_type_c_changed = (
        old_solution != (task.cach_giai_quyet or "") or
        old_approval != (task.duyet_tuan or "Chưa duyệt")
    )
    
    recalculate_budgets(db)
    
    if is_type_a_changed:
        log_details = f"Cập nhật thông tin công việc {task.stt} - {task.ten_cong_viec}. Tiến độ: {task.tien_do}%, Trạng thái: {task.trang_thai}."
        log_action(db, username, "Chỉnh sửa công việc", log_details)
        
    if is_type_b_changed:
        plan_text = task.ke_hoach_tuan if task.ke_hoach_tuan else "-"
        result_text = task.ket_qua_tuan if task.ket_qua_tuan else "-"
        log_details = f"Báo cáo kế hoạch/kết quả tuần cho công việc {task.stt} - {task.ten_cong_viec}. Kế hoạch: \"{plan_text}\", Kết quả: \"{result_text}\"."
        log_action(db, username, "Báo cáo tuần", log_details)
        
    if is_type_c_changed:
        solution_text = task.cach_giai_quyet if task.cach_giai_quyet else "-"
        approval_text = task.duyet_tuan if task.duyet_tuan else "Chưa duyệt"
        log_details = f"CBQL phê duyệt báo cáo tuần cho công việc {task.stt} - {task.ten_cong_viec}. Trạng thái duyệt: \"{approval_text}\", Giải pháp: \"{solution_text}\"."
        log_action(db, username, "Phê duyệt tuần", log_details)
        
    db.commit()
    
    update_msg = {
        "type": "task_update",
        "task_id": task.id,
        "stt": task.stt,
        "ma_ngan_sach": task.ma_ngan_sach,
        "ten_cong_viec": task.ten_cong_viec,
        "phong_ban_thuc_hien": task.phong_ban_thuc_hien,
        "ho_so_dau_ra": task.ho_so_dau_ra,
        "thoi_han_hoan_thanh": task.thoi_han_hoan_thanh,
        "tien_do": task.tien_do,
        "trang_thai": task.trang_thai,
        "ke_hoach_tuan": task.ke_hoach_tuan,
        "ket_qua_tuan": task.ket_qua_tuan,
        "vuong_mac_tuan": task.vuong_mac_tuan,
        "cach_giai_quyet": task.cach_giai_quyet,
        "duyet_tuan": task.duyet_tuan,
        "ngay_khoi_tao": task.ngay_khoi_tao,
        "cong_trinh": task.cong_trinh,
        "doi_tac": task.doi_tac,
        "so_dien_thoai": task.so_dien_thoai,
        "ngay_bat_dau": task.ngay_bat_dau,
        "ngay_ket_thuc": task.ngay_ket_thuc,
        "gia_han_den_ngay": task.gia_han_den_ngay,
        "thoi_gian_bao_hanh": task.thoi_gian_bao_hanh,
        "mo_ta": task.mo_ta,
        "dieu_khoan": task.dieu_khoan,
        "nguoi_phu_trach": task.nguoi_phu_trach,
        "nguoi_bao_cao": task.nguoi_bao_cao,
        "nguoi_duyet": task.nguoi_duyet,
        "gia_tri_quyet_toan": task.gia_tri_quyet_toan,
        "da_nghiem_thu": task.da_nghiem_thu,
        "da_thanh_toan": task.da_thanh_toan,
        "tam_ung": task.tam_ung,
        "da_thu_hoi_tam_ung": task.da_thu_hoi_tam_ung,
        "weekly_reports_json": task.weekly_reports_json,
        "budget": {
            "ngan_sach_tong": task.budget.ngan_sach_tong if task.budget else 0.0,
            "is_locked": task.budget.is_locked if task.budget else False
        }
    }
    await manager.broadcast(update_msg)
    
    return {"status": "success", "task": update_msg}

@app.delete("/api/tasks/{task_id}")
async def delete_task_route(
    task_id: int, 
    username: Optional[str] = None,
    db: Session = Depends(database.get_db)
):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Không tìm thấy công việc")
        
    user = None
    if username:
        user = db.query(models.User).filter(models.User.username == username).first()
        
    task_level = task.stt.count('.') + 1
    
    # Rule 2: Admin or PM only for Level 1 & 2 tasks
    # Rule 1: Only Admin can delete Level 1 & 2 tasks
    if task_level <= 2:
        if not user or user.role != "Admin":
            raise HTTPException(
                status_code=403,
                detail="Chỉ có Admin mới được phép xóa công việc cấp 1, cấp 2."
            )
            
    # Rule 2 & 3: Level >= 3 tasks
    if task_level >= 3:
        if not user:
            raise HTTPException(status_code=403, detail="Yêu cầu đăng nhập để thực hiện xóa công việc.")
            
        user_dept = str(user.phong_ban).upper().strip()
        task_dept = str(task.phong_ban_thuc_hien).upper().strip()
        
        # Non-admins must belong to the department of the task
        if user.role != "Admin" and user_dept != task_dept and user_dept != "ALL":
            raise HTTPException(
                status_code=403,
                detail=f"Quyền hạn bị từ chối: Phòng ban của bạn ({user.phong_ban}) không trùng khớp với phòng thực hiện công việc này ({task.phong_ban_thuc_hien})."
            )
            
        # Level 3 tasks: only Admin or Trưởng phòng (TruongPhong/CBQL) can delete
        if task_level == 3 and user.role != "Admin":
            if user.role not in ("TruongPhong", "CBQL", "PM"):
                raise HTTPException(
                    status_code=403,
                    detail="Chỉ có Admin hoặc Trưởng phòng mới được phép xóa công việc cấp 3."
                )
            
        # Level 4 and 5 tasks: Admin, Trưởng phòng, or NhanVien of the same department can delete (all are allowed if department matches)
        
    stt = task.stt
    children = db.query(models.Task).filter(
        (models.Task.stt == stt) | (models.Task.stt.like(f"{stt}.%"))
    ).all()
    
    deleted_ids = [c.id for c in children]
    for c in children:
        db.delete(c)
    db.flush()
    
    recalculate_budgets(db)
    log_action(db, username, "Xóa công việc", f"Xóa công việc {stt} - {task.ten_cong_viec} và {len(children)-1} công việc con.")
    db.commit()
    
    await manager.broadcast({
        "type": "task_delete",
        "task_ids": deleted_ids
    })
    
    return {"status": "success", "deleted_count": len(children)}

# Serve Frontend static files
# We will create the static directory if not exists
os.makedirs("./static", exist_ok=True)
app.mount("/", StaticFiles(directory="./static", html=True), name="static")

import asyncio

async def monitor_observability_loop():
    await asyncio.sleep(5)  # Chờ ứng dụng khởi động hoàn toàn
    while True:
        db = database.SessionLocal()
        try:
            # 1. Kiểm tra trễ hạn công việc
            # Thời điểm hiện tại là Tháng 7/2026. Do đó các kỳ hạn trong quá khứ nếu
            # có ngân sách kế hoạch lớn hơn 0 mà tiến độ < 100% thì được coi là trễ hạn.
            delayed_count = 0
            tasks = db.query(models.Task).all()
            for t in tasks:
                if t.tien_do < 100.0 and t.budget:
                    b = t.budget
                    had_past_plan = (
                        b.thang_01_2025 > 0 or b.thang_02_2025 > 0 or
                        b.thang_03_2025 > 0 or b.thang_04_2025 > 0 or
                        b.quy_1_2026 > 0 or b.thang_05_2026 > 0 or
                        b.thang_06_2026 > 0
                    )
                    if had_past_plan:
                        delayed_count += 1

            # 2. Kiểm tra lệch ngân sách (> 10% lệch so với kế hoạch tháng gần nhất - Tháng 6/2026)
            plan_june_2026 = sum(t.budget.thang_06_2026 for t in tasks if t.budget)
            
            # Tính thực chi Tháng 6/2026
            spendings_june_2026 = db.query(models.ActualSpending).filter(
                models.ActualSpending.trang_thai_duyet == "Approved",
                models.ActualSpending.ngay_chi >= datetime.date(2026, 6, 1),
                models.ActualSpending.ngay_chi <= datetime.date(2026, 6, 30)
            ).all()
            actual_june_2026 = sum(s.so_tien_chi for s in spendings_june_2026)

            if plan_june_2026 > 0:
                deviation = abs(actual_june_2026 - plan_june_2026) / plan_june_2026
                if deviation > 0.10:
                    alert_msg = (
                        f"🚨 CẢNH BÁO LỆCH CHI TIÊU THÁNG 06/2026: Kế hoạch {plan_june_2026:,.2f} Trđ, "
                        f"Thực tế giải ngân {actual_june_2026:,.2f} Trđ (Lệch {deviation*100:.1f}% - vượt hạn mức 10%)."
                    )
                    print(f"[OBSERVABILITY] {alert_msg}")
                    await manager.broadcast({
                        "type": "red_alert",
                        "message": alert_msg,
                        "ma_ngan_sach": "THANG_06_2026"
                    })

            if delayed_count > 0:
                summary = f"⚠️ CẢNH BÁO TIẾN ĐỘ: Phát hiện {delayed_count} công việc trong quá khứ đã quá thời hạn phân bổ ngân sách nhưng chưa đạt tiến độ 100%."
                print(f"[OBSERVABILITY] {summary}")
                await manager.broadcast({
                    "type": "red_alert",
                    "message": summary,
                    "ma_ngan_sach": "TIEN_DO"
                })

        except Exception as e:
            print(f"Lỗi trong vòng lặp Observability: {e}")
        finally:
            db.close()
        
        await asyncio.sleep(20)  # Chạy quét sau mỗi 20 giây cho demo

@app.on_event("startup")
async def startup_event():
    asyncio.create_task(monitor_observability_loop())
